import json
import requests
from openpyxl import load_workbook
from openpyxl.styles import Color
from io import BytesIO

def handler(event, context):
    """Download and parse Excel file to extract contract template structure"""
    
    # URL of the Excel file
    url = "https://cdn.poehali.dev/projects/19fc2d54-8db4-4df4-b970-568be02f70ef/bucket/СИГ%20%2B%20ФМ%20Д-З%20Мск-Ижевск%20(Шабердино)%20лук%20на%2020.12.2025%20вод.%20Шильков.xlsx"
    
    try:
        # Download the Excel file
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        # Load workbook from bytes
        wb = load_workbook(BytesIO(response.content))
        
        # Get the active sheet
        sheet = wb.active
        sheet_name = sheet.title
        
        # Initialize result structure
        result = {
            "sheet_name": sheet_name,
            "dimensions": {
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "min_row": sheet.min_row,
                "min_column": sheet.min_column
            },
            "merged_cells": [],
            "cells": [],
            "red_cells": [],
            "tables": []
        }
        
        # Extract merged cell ranges
        for merged_range in sheet.merged_cells.ranges:
            result["merged_cells"].append({
                "range": str(merged_range),
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col
            })
        
        # Extract all cell data
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                                   min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value is not None or cell.fill or cell.font:
                    cell_info = {
                        "coordinate": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "value": str(cell.value) if cell.value is not None else "",
                        "data_type": cell.data_type,
                        "font": {},
                        "fill": {},
                        "alignment": {},
                        "border": {}
                    }
                    
                    # Extract font information
                    if cell.font:
                        cell_info["font"] = {
                            "name": cell.font.name,
                            "size": cell.font.size,
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "color": cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
                        }
                        
                        # Check if font color is red
                        if cell.font.color and hasattr(cell.font.color, 'rgb'):
                            color_rgb = cell.font.color.rgb
                            if color_rgb:
                                # Check for red colors (FF0000 or variations)
                                if isinstance(color_rgb, str) and len(color_rgb) >= 6:
                                    red_value = int(color_rgb[-6:-4], 16)
                                    green_value = int(color_rgb[-4:-2], 16)
                                    blue_value = int(color_rgb[-2:], 16)
                                    
                                    # Red if red channel is high and green/blue are low
                                    if red_value > 200 and green_value < 100 and blue_value < 100:
                                        result["red_cells"].append({
                                            "coordinate": cell.coordinate,
                                            "row": cell.row,
                                            "column": cell.column,
                                            "value": str(cell.value) if cell.value else "",
                                            "color_rgb": color_rgb
                                        })
                    
                    # Extract fill information
                    if cell.fill:
                        cell_info["fill"] = {
                            "patternType": cell.fill.patternType,
                            "fgColor": cell.fill.fgColor.rgb if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None,
                            "bgColor": cell.fill.bgColor.rgb if cell.fill.bgColor and hasattr(cell.fill.bgColor, 'rgb') else None
                        }
                    
                    # Extract alignment
                    if cell.alignment:
                        cell_info["alignment"] = {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text
                        }
                    
                    # Extract border info
                    if cell.border:
                        cell_info["border"] = {
                            "left": str(cell.border.left.style) if cell.border.left else None,
                            "right": str(cell.border.right.style) if cell.border.right else None,
                            "top": str(cell.border.top.style) if cell.border.top else None,
                            "bottom": str(cell.border.bottom.style) if cell.border.bottom else None
                        }
                    
                    result["cells"].append(cell_info)
        
        # Create summary
        result["summary"] = {
            "total_cells_with_data": len(result["cells"]),
            "total_merged_ranges": len(result["merged_cells"]),
            "total_red_cells": len(result["red_cells"]),
            "red_cells_list": [f"{cell['coordinate']}: {cell['value']}" for cell in result["red_cells"]]
        }
        
        return {
            "statusCode": 200,
            "headers": {
                "Content-Type": "application/json",
                "Access-Control-Allow-Origin": "*"
            },
            "body": json.dumps(result, ensure_ascii=False, indent=2)
        }
        
    except Exception as e:
        return {
            "statusCode": 500,
            "headers": {
                "Content-Type": "application/json",
                "Access-Control-Allow-Origin": "*"
            },
            "body": json.dumps({
                "error": str(e),
                "type": type(e).__name__
            }, ensure_ascii=False)
        }