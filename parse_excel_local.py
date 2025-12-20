#!/usr/bin/env python3
"""
Excel Contract Template Parser
Downloads and analyzes Excel file to extract structure, formatting, and variable fields
"""

import json
import requests
from openpyxl import load_workbook
from io import BytesIO

def parse_excel_contract():
    """Download and parse Excel file to extract contract template structure"""
    
    # URL of the Excel file
    url = "https://cdn.poehali.dev/projects/19fc2d54-8db4-4df4-b970-568be02f70ef/bucket/СИГ%20%2B%20ФМ%20Д-З%20Мск-Ижевск%20(Шабердино)%20лук%20на%2020.12.2025%20вод.%20Шильков.xlsx"
    
    print("Downloading Excel file...")
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    print(f"✓ Downloaded {len(response.content)} bytes\n")
    
    # Load workbook from bytes
    print("Loading workbook...")
    wb = load_workbook(BytesIO(response.content))
    sheet = wb.active
    print(f"✓ Loaded sheet: '{sheet.title}'\n")
    
    # Initialize result structure
    result = {
        "sheet_name": sheet.title,
        "dimensions": {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "min_row": sheet.min_row,
            "min_column": sheet.min_column
        },
        "merged_cells": [],
        "cells": [],
        "red_cells": [],
    }
    
    print(f"Document Dimensions:")
    print(f"  Rows: {sheet.min_row} to {sheet.max_row}")
    print(f"  Columns: {sheet.min_column} to {sheet.max_column}\n")
    
    # Extract merged cell ranges
    print("Extracting merged cells...")
    for merged_range in sheet.merged_cells.ranges:
        result["merged_cells"].append({
            "range": str(merged_range),
            "min_row": merged_range.min_row,
            "max_row": merged_range.max_row,
            "min_col": merged_range.min_col,
            "max_col": merged_range.max_col
        })
    print(f"✓ Found {len(result['merged_cells'])} merged cell ranges\n")
    
    # Extract all cell data
    print("Extracting cell data and formatting...")
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                               min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None or cell.fill or cell.font:
                cell_info = {
                    "coordinate": cell.coordinate,
                    "row": cell.row,
                    "column": cell.column,
                    "value": str(cell.value) if cell.value is not None else "",
                    "data_type": cell.data_type,
                    "font": {},
                    "fill": {},
                    "alignment": {},
                    "border": {}
                }
                
                # Extract font information
                if cell.font:
                    cell_info["font"] = {
                        "name": cell.font.name,
                        "size": cell.font.size,
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "underline": cell.font.underline,
                        "color": None
                    }
                    
                    # Extract font color
                    if cell.font.color and hasattr(cell.font.color, 'rgb'):
                        color_rgb = cell.font.color.rgb
                        cell_info["font"]["color"] = color_rgb
                        
                        # Check if font color is red
                        if color_rgb and isinstance(color_rgb, str) and len(color_rgb) >= 6:
                            red_value = int(color_rgb[-6:-4], 16)
                            green_value = int(color_rgb[-4:-2], 16)
                            blue_value = int(color_rgb[-2:], 16)
                            
                            # Red if red channel is high and green/blue are low
                            if red_value > 200 and green_value < 100 and blue_value < 100:
                                result["red_cells"].append({
                                    "coordinate": cell.coordinate,
                                    "row": cell.row,
                                    "column": cell.column,
                                    "value": str(cell.value) if cell.value else "",
                                    "color_rgb": color_rgb
                                })
                
                # Extract fill information
                if cell.fill:
                    cell_info["fill"] = {
                        "patternType": cell.fill.patternType,
                        "fgColor": cell.fill.fgColor.rgb if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None,
                        "bgColor": cell.fill.bgColor.rgb if cell.fill.bgColor and hasattr(cell.fill.bgColor, 'rgb') else None
                    }
                
                # Extract alignment
                if cell.alignment:
                    cell_info["alignment"] = {
                        "horizontal": cell.alignment.horizontal,
                        "vertical": cell.alignment.vertical,
                        "wrap_text": cell.alignment.wrap_text,
                        "text_rotation": cell.alignment.text_rotation
                    }
                
                # Extract border info
                if cell.border:
                    cell_info["border"] = {
                        "left": str(cell.border.left.style) if cell.border.left else None,
                        "right": str(cell.border.right.style) if cell.border.right else None,
                        "top": str(cell.border.top.style) if cell.border.top else None,
                        "bottom": str(cell.border.bottom.style) if cell.border.bottom else None
                    }
                
                result["cells"].append(cell_info)
    
    print(f"✓ Extracted {len(result['cells'])} cells with data\n")
    
    # Create summary
    result["summary"] = {
        "total_cells_with_data": len(result["cells"]),
        "total_merged_ranges": len(result["merged_cells"]),
        "total_red_cells": len(result["red_cells"]),
        "red_cells_list": [f"{cell['coordinate']}: {cell['value']}" for cell in result["red_cells"]]
    }
    
    # Print summary
    print("=" * 60)
    print("ANALYSIS SUMMARY")
    print("=" * 60)
    print(f"Sheet Name: {result['sheet_name']}")
    print(f"Total Rows Used: {result['dimensions']['max_row']}")
    print(f"Total Columns Used: {result['dimensions']['max_column']}")
    print(f"Total Cells with Data: {result['summary']['total_cells_with_data']}")
    print(f"Total Merged Cell Ranges: {result['summary']['total_merged_ranges']}")
    print(f"Total RED Cells (Variables): {result['summary']['total_red_cells']}")
    print()
    
    if result["red_cells"]:
        print("RED CELLS (Variable Fields):")
        print("-" * 60)
        for red_cell in result["red_cells"]:
            print(f"  {red_cell['coordinate']:<6} (Row {red_cell['row']:>3}, Col {red_cell['column']:>2}): {red_cell['value']}")
        print()
    
    if result["merged_cells"]:
        print(f"MERGED CELL RANGES (First 20):")
        print("-" * 60)
        for i, merged in enumerate(result["merged_cells"][:20]):
            print(f"  {merged['range']:<15} rows {merged['min_row']:>3}-{merged['max_row']:<3} cols {merged['min_col']:>2}-{merged['max_col']:<2}")
        if len(result["merged_cells"]) > 20:
            print(f"  ... and {len(result['merged_cells']) - 20} more")
        print()
    
    # Save to JSON file
    output_file = "contract_template_analysis.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"✓ Full analysis saved to: {output_file}")
    print("=" * 60)
    
    return result

if __name__ == "__main__":
    try:
        parse_excel_contract()
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
